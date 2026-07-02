import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load & filter ──────────────────────────────────────────────────────────────
DATA_DIR = "/home/juan_esteban_correa/qargo-data/data"
FILES = ["DDBB_Mar_26.csv", "DDBB_Apr_26.csv", "DDBB_May_26.csv"]

dfs = [pd.read_csv(f"{DATA_DIR}/{f}") for f in FILES]
raw = pd.concat(dfs, ignore_index=True)

gp = raw[
    raw["Location"].str.contains("Grand Prairie", na=False) &
    (raw["Voided"] == False)
].copy()

def parse_currency(s):
    s = str(s).replace("$", "").replace(",", "").strip()
    if s.startswith("(") and s.endswith(")"):
        return -float(s[1:-1])
    return float(s) if s not in ("", "nan") else 0.0

gp["Net Sales"] = gp["Net Sales"].apply(parse_currency)
gp["date"] = pd.to_datetime(gp["Closed Date/Time"])
gp["day"]  = gp["date"].dt.date

# Drive-Thru vs everything else
gp["is_drive_thru"] = gp["Destination"] == "DRIVE THRU"
gp["is_non_drive"]  = gp["Destination"] != "DRIVE THRU"
gp["is_item"]       = gp["Is Modifier"] == False   # exclude modifiers for item count

# ── Verification: channel breakdown ───────────────────────────────────────────
dest_breakdown = gp.groupby("Destination")["Net Sales"].sum().sort_values(ascending=False)
print("=== Destination breakdown (Net Sales) ===")
print(dest_breakdown.to_string())
print(f"\nDrive-Thru total : ${gp.loc[gp['is_drive_thru'], 'Net Sales'].sum():,.2f}")
print(f"Non-Drive total  : ${gp.loc[gp['is_non_drive'], 'Net Sales'].sum():,.2f}")
print(f"Grand total      : ${gp['Net Sales'].sum():,.2f}")
print(f"Total items      : {(gp['is_item']).sum():,}")

# ── Aggregation helpers ───────────────────────────────────────────────────────
def sum_sales(x, flag_col):
    return x[gp.loc[x.index, flag_col]].sum()

def count_items(x, flag_col):
    mask = gp.loc[x.index, flag_col] & gp.loc[x.index, "is_item"]
    return int(mask.sum())

# ── Daily aggregation ─────────────────────────────────────────────────────────
daily = (
    gp.groupby("day")
      .agg(
          drv_sales  = ("Net Sales", lambda x: sum_sales(x, "is_drive_thru")),
          drv_items  = ("Net Sales", lambda x: count_items(x, "is_drive_thru")),
          non_sales  = ("Net Sales", lambda x: sum_sales(x, "is_non_drive")),
          non_items  = ("Net Sales", lambda x: count_items(x, "is_non_drive")),
          total      = ("Net Sales", "sum"),
          tot_items  = ("is_item",   "sum"),
      )
      .reset_index()
)
daily.columns = [
    "Date",
    "Drive-Thru ($)", "Drive-Thru Items",
    "At-Store & Delivery ($)", "At-Store & Delivery Items",
    "Total Net Sales ($)", "Total Items Sold",
]
daily["Date"] = pd.to_datetime(daily["Date"])
daily = daily.sort_values("Date")

# ── Weekly aggregation ────────────────────────────────────────────────────────
gp["week_start"] = gp["date"].dt.to_period("W").apply(lambda p: p.start_time.date())

weekly = (
    gp.groupby("week_start")
      .agg(
          drv_sales  = ("Net Sales", lambda x: sum_sales(x, "is_drive_thru")),
          drv_items  = ("Net Sales", lambda x: count_items(x, "is_drive_thru")),
          non_sales  = ("Net Sales", lambda x: sum_sales(x, "is_non_drive")),
          non_items  = ("Net Sales", lambda x: count_items(x, "is_non_drive")),
          total      = ("Net Sales", "sum"),
          tot_items  = ("is_item",   "sum"),
      )
      .reset_index()
)
weekly.columns = [
    "Week Starting (Mon)",
    "Drive-Thru ($)", "Drive-Thru Items",
    "At-Store & Delivery ($)", "At-Store & Delivery Items",
    "Total Net Sales ($)", "Total Items Sold",
]
weekly["Week Starting (Mon)"] = pd.to_datetime(weekly["Week Starting (Mon)"])
weekly = weekly.sort_values("Week Starting (Mon)")

# ── Cross-checks ──────────────────────────────────────────────────────────────
assert abs(daily["Drive-Thru ($)"].sum() - weekly["Drive-Thru ($)"].sum()) < 0.01, "Drive-Thru $ mismatch"
assert abs(daily["Total Net Sales ($)"].sum() - gp["Net Sales"].sum()) < 0.01, "Total $ mismatch"
assert daily["Total Items Sold"].sum() == weekly["Total Items Sold"].sum(), "Items mismatch daily vs weekly"
assert daily["Total Items Sold"].sum() == gp["is_item"].sum(), "Items mismatch vs raw"
print("\n[OK] Cross-checks passed.")

# ── Excel ──────────────────────────────────────────────────────────────────────
OUT = "/home/juan_esteban_correa/qargo-data/grand_prairie_channel_sales_Q2_2026.xlsx"

NAV   = "1F3864"
BLUE  = "2E75B6"
GREEN = "2E8B57"
GRAY  = "404040"
LBLUE = "EDF4FB"
WHITE = "FFFFFF"

def side():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(ws, df, title):
    ws.title = title
    ncols = len(df.columns)

    # col index -> header color, data format
    # cols: Date | DRV$ | DRV# | NON$ | NON# | TOT$ | TOT#
    HDR_COLORS = [NAV, BLUE, BLUE, GREEN, GREEN, GRAY, GRAY]
    COL_FMTS   = [
        "MMM DD, YYYY",
        '"$"#,##0.00', "#,##0",
        '"$"#,##0.00', "#,##0",
        '"$"#,##0.00', "#,##0",
    ]

    # Title
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = f"Grand Prairie, TX  |  Sales by Channel  |  {title}  |  Mar – May 2026"
    t.font = Font(bold=True, size=12, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAV)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Column headers
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_COLORS[ci - 1])
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = side()
    ws.row_dimensions[2].height = 32

    # Data
    for ri, row in enumerate(df.itertuples(index=False), 3):
        bg = PatternFill("solid", fgColor=LBLUE if ri % 2 == 0 else WHITE)
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = side()
            c.fill = bg
            c.alignment = Alignment(horizontal="center")
            c.number_format = COL_FMTS[ci - 1]

    # Totals
    tr = len(df) + 3
    c0 = ws.cell(row=tr, column=1, value="TOTAL")
    c0.font = Font(bold=True, color=WHITE)
    c0.fill = PatternFill("solid", fgColor=NAV)
    c0.alignment = Alignment(horizontal="center")
    c0.border = side()
    for ci in range(2, ncols + 1):
        cl = get_column_letter(ci)
        c = ws.cell(row=tr, column=ci)
        c.value = f"=SUM({cl}3:{cl}{tr-1})"
        c.number_format = COL_FMTS[ci - 1]
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAV)
        c.alignment = Alignment(horizontal="center")
        c.border = side()

    # Widths
    ws.column_dimensions["A"].width = 22
    col_widths = [22, 20, 18, 24, 22, 20, 18]
    for ci in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
    ws.freeze_panes = "A3"


wb = openpyxl.Workbook()
wb.remove(wb.active)

write_sheet(wb.create_sheet(), daily, "Day by Day")
write_sheet(wb.create_sheet(), weekly, "Week by Week")

# ── Verification sheet ────────────────────────────────────────────────────────
vs = wb.create_sheet("Verification")
vs["A1"] = "Data Verification"
vs["A1"].font = Font(bold=True, size=12)
drv = gp.loc[gp["is_drive_thru"], "Net Sales"].sum()
non = gp.loc[gp["is_non_drive"], "Net Sales"].sum()
tot = gp["Net Sales"].sum()
drv_items = int((gp["is_drive_thru"] & gp["is_item"]).sum())
non_items = int((gp["is_non_drive"] & gp["is_item"]).sum())
tot_items = int(gp["is_item"].sum())

rows = [
    ("Metric", "Value"),
    ("Source files", ", ".join(FILES)),
    ("Store", "Qargo Coffee Grand Prairie, TX"),
    ("Voided excluded", "Yes"),
    ("Modifiers excluded from item count", "Yes (Is Modifier = True)"),
    ("Total rows (non-voided)", f"{len(gp):,}"),
    ("Date range", f"{gp['day'].min()} to {gp['day'].max()}"),
    ("", ""),
    ("Drive-Thru Net Sales", f"${drv:,.2f}"),
    ("Drive-Thru Items Sold", f"{drv_items:,}"),
    ("At-Store & Delivery Net Sales", f"${non:,.2f}"),
    ("At-Store & Delivery Items Sold", f"{non_items:,}"),
    ("Total Net Sales", f"${tot:,.2f}"),
    ("Total Items Sold", f"{tot_items:,}"),
    ("Drive-Thru % of Total Sales", f"{drv/tot*100:.1f}%"),
    ("At-Store % of Total Sales", f"{non/tot*100:.1f}%"),
    ("", ""),
    ("Destination breakdown (Net Sales)", ""),
]
for d, s in dest_breakdown.items():
    rows.append((f"  · {d}", f"${s:,.2f}"))

for ri, (k, v) in enumerate(rows, 1):
    vs.cell(row=ri, column=1, value=k).font = Font(bold=(ri == 1))
    vs.cell(row=ri, column=2, value=v)
vs.column_dimensions["A"].width = 38
vs.column_dimensions["B"].width = 30

wb.save(OUT)
print(f"Saved → {OUT}")
