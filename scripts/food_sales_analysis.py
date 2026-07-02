import duckdb
import pandas as pd
import glob
import re
import warnings
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

warnings.filterwarnings("ignore")

PRODUCTS = [
    "Cake Red Velvet",
    "Cake Chocolate Temptation",
    "Cheesecake Pistachio",
    "Macarons French",
    "Cake Carrot",
    "Cake Cappuccino",
    "Cheesecake Dulce de Leche",
    "Cheesecake Strawberry",
    "Tiramisu With Ladyfingers",
    "Tiramisu",
    "Cheesecake Chocolate Ganache",
    "Cake Torta Nocciola",
    "Tart Mixed Berry",
    "Cake Ricotta and Pistachio",
    "Cake Chocolate Fondant",
    "Cake Tres Leches",
    "Cake Red Velvet Mini Gluten Free",
    "Cheesecake Brownie",
]

# ── DDBB item classifier ──────────────────────────────────────────────────────
def classify_ddbb(s):
    s = s.lower().strip()

    cheese_kw  = ["cheesecake","cheese cake","chees cake","chesscake","chescake",
                  "chssecake","chz cake","chez cake","cheeze cake","cheescake",
                  "chesecake","cheeaecake","cheasecake","cheesecakw","che cake"]
    pist_kw    = ["pistachio","pistashio","pistacio","piatachio","pitachio","pstachio"]
    ricotta_kw = ["ricotta","ricota","riotta","icotta"]
    brownie_kw = ["brownie","browni","brawni","brwni",".brawni"]
    tiram_kw   = ["tiramisu","taramisu","tiramiso","tirmisu"]

    is_cheese  = any(k in s for k in cheese_kw)
    is_pist    = any(k in s for k in pist_kw)
    is_ricotta = any(k in s for k in ricotta_kw)
    is_brownie = any(k in s for k in brownie_kw)
    is_tiram   = any(k in s for k in tiram_kw)

    # mini GF red velvet
    if ("velvet" in s or "revelvet" in s) and ("mini" in s or "gluten" in s):
        return "Cake Red Velvet Mini Gluten Free"

    # red velvet standard
    if any(k in s for k in ["red velvet","re velvet","red velt","redvelvet","rv cake"]):
        if "mini" not in s and "cupcake" not in s and "cup cake" not in s:
            return "Cake Red Velvet"

    # chocolate temptation
    if "temptation" in s:
        return "Cake Chocolate Temptation"

    # ricotta + pistachio (check before plain pistachio cheesecake)
    if is_pist and is_ricotta:
        return "Cake Ricotta and Pistachio"
    if is_ricotta and ("cake" in s or "cak" in s):
        return "Cake Ricotta and Pistachio"

    # pistachio cheesecake (not ricotta)
    if is_pist and is_cheese and not is_ricotta:
        return "Cheesecake Pistachio"

    # macarons
    if "macaron" in s:
        return "Macarons French"

    # carrot cake
    if any(k in s for k in ["carrot","carot","carott"]) and any(k in s for k in ["cake","cak"]):
        return "Cake Carrot"

    # cappuccino cake
    if "cappuccino" in s and any(k in s for k in ["cake","cak"]):
        return "Cake Cappuccino"

    # dulce de leche cheesecake
    if "dulce" in s and (is_cheese or "leche" in s):
        return "Cheesecake Dulce de Leche"

    # strawberry cheesecake
    straw = ["strawberry","straberry","strabery","strwberry","strabrry","stwberry","trawberry"]
    if any(k in s for k in straw) and is_cheese:
        return "Cheesecake Strawberry"
    if "fragola" in s and is_cheese:
        return "Cheesecake Strawberry"
    if "2 strawberry" in s and "cake" in s:
        return "Cheesecake Strawberry"

    # tiramisu with ladyfingers (before plain tiramisu)
    if is_tiram and "ladyfinger" in s:
        return "Tiramisu With Ladyfingers"

    # tiramisu plain
    if is_tiram and "latte" not in s and "ladyfinger" not in s:
        return "Tiramisu"

    # ganache cheesecake
    if "ganache" in s and is_cheese:
        return "Cheesecake Chocolate Ganache"
    if "choc gan" in s:
        return "Cheesecake Chocolate Ganache"

    # torta nocciola
    if "nocciola" in s:
        return "Cake Torta Nocciola"

    # mixed berry tart
    if ("berry" in s or "bery" in s) and ("tart" in s or "trt" in s or "tar" == s[-3:]):
        return "Tart Mixed Berry"

    # chocolate fondant
    if "fondant" in s:
        return "Cake Chocolate Fondant"

    # tres leches
    if "tres leche" in s or "tres lech" in s:
        return "Cake Tres Leches"

    # brownie cheesecake
    if is_brownie and is_cheese:
        return "Cheesecake Brownie"

    return None


# ── Berkeley item classifier ──────────────────────────────────────────────────
BERK_MAP = {
    "torta nocciola":                "Cake Torta Nocciola",
    "red velvet cake":               "Cake Red Velvet",
    "mini gf red velvet cake":       "Cake Red Velvet Mini Gluten Free",
    "mini gluten-free red velvet":   "Cake Red Velvet Mini Gluten Free",
    "tres leches cake":              "Cake Tres Leches",
    "chocolate ganache cheesecake":  "Cheesecake Chocolate Ganache",
    "macarons french collection":    "Macarons French",
    "french macarons":               "Macarons French",
    "macaron":                       "Macarons French",
    "strawberry cheesecake":         "Cheesecake Strawberry",
    "tiramisu individual":           "Tiramisu",
    "tiramisu big ladyfinger":       "Tiramisu With Ladyfingers",
    "tiramisu with ladyfingers":     "Tiramisu With Ladyfingers",
    "mixed berry tart":              "Tart Mixed Berry",
    "chocolate temptation cake":     "Cake Chocolate Temptation",
    "cheesecake pistachio":          "Cheesecake Pistachio",
    "pistachio cheesecake":          "Cheesecake Pistachio",
    "cappuccino cake":               "Cake Cappuccino",
    "carrot cake":                   "Cake Carrot",
    "dulce de leche cheesecake":     "Cheesecake Dulce de Leche",
    "chocolate fondant cake":        "Cake Chocolate Fondant",
    "cheesecake brownie":            "Cheesecake Brownie",
    "brownie cheesecake":            "Cheesecake Brownie",
    "ricotta & pistachio cake":      "Cake Ricotta and Pistachio",
    "ricotta and pistachio cake":    "Cake Ricotta and Pistachio",
}

def classify_berkeley(s):
    low = s.lower().strip()
    if low in BERK_MAP:
        return BERK_MAP[low]
    # fallback to same logic as DDBB
    return classify_ddbb(low)


# ── Load DDBB data ────────────────────────────────────────────────────────────
print("Loading DDBB files...")
con = duckdb.connect()
ddbb_raw = con.execute("""
    SELECT
        "Location"       AS location,
        "Closed Date/Time" AS dt_raw,
        "Item Name"      AS item_name,
        "Net Sales"      AS net_sales_raw,
        "Is Modifier"    AS is_modifier,
        "Voided"         AS voided
    FROM read_csv_auto('data/DDBB_*.csv', union_by_name=true, header=true)
""").df()

ddbb = ddbb_raw[
    (ddbb_raw["voided"].astype(str).str.strip().str.lower() == "false") &
    (ddbb_raw["is_modifier"].astype(str).str.strip().str.lower() == "false")
].copy()

ddbb["net_sales"] = (
    ddbb["net_sales_raw"]
    .astype(str)
    .str.replace("$", "", regex=False)
    .str.replace(",", "", regex=False)
    .str.strip()
)
ddbb["net_sales"] = pd.to_numeric(ddbb["net_sales"], errors="coerce").fillna(0)
ddbb["date"] = pd.to_datetime(ddbb["dt_raw"], format="mixed", errors="coerce")
ddbb = ddbb.dropna(subset=["date"])
ddbb["product"] = ddbb["item_name"].apply(classify_ddbb)
ddbb = ddbb[ddbb["product"].notna()].copy()
ddbb["source"] = "DDBB"
print(f"  DDBB rows matched: {len(ddbb)}")

# ── Load Berkeley data ────────────────────────────────────────────────────────
print("Loading Berkeley files...")
berk_files = glob.glob("data/qargocoffee*.csv")
berk_dfs = [pd.read_csv(f, sep=";", encoding="latin-1", on_bad_lines="skip") for f in berk_files]
berk_raw = pd.concat(berk_dfs, ignore_index=True)

berk = berk_raw[berk_raw["Type"] == "SALE"].copy()
berk["date"]     = pd.to_datetime(berk["Date"], format="mixed", errors="coerce")
berk["location"] = "Qargo Coffee Berkeley, CA"
berk["net_sales"] = pd.to_numeric(berk["FinalPrice"], errors="coerce").fillna(0)
berk["product"]  = berk["Item"].apply(classify_berkeley)
berk = berk[berk["product"].notna()].copy()
berk["source"] = "Berkeley"
print(f"  Berkeley rows matched: {len(berk)}")

# ── Combine ───────────────────────────────────────────────────────────────────
all_data = pd.concat(
    [
        ddbb[["date","location","product","net_sales","item_name","source"]],
        berk[["date","location","product","net_sales"]].assign(
            item_name=berk["Item"], source=berk["source"]
        ),
    ],
    ignore_index=True,
)
all_data["month"]      = all_data["date"].dt.to_period("M")
all_data["month_str"]  = all_data["date"].dt.strftime("%Y-%m")
all_data["month_label"] = all_data["date"].dt.strftime("%b %Y")

print(f"  Total matched rows: {len(all_data)}")
print("\nProduct counts:")
print(all_data.groupby("product")["net_sales"].agg(["sum","count"]).round(2).to_string())

# ── Aggregated: month × product ───────────────────────────────────────────────
agg = (
    all_data.groupby(["month_str","product"])
    .agg(net_sales=("net_sales","sum"), units=("net_sales","count"))
    .reset_index()
)

# Pivot for Net Sales
months = sorted(agg["month_str"].unique())
pivot_sales = agg.pivot(index="month_str", columns="product", values="net_sales").reindex(index=months, columns=PRODUCTS).fillna(0)
pivot_units = agg.pivot(index="month_str", columns="product", values="units").reindex(index=months, columns=PRODUCTS).fillna(0).astype(int)

# Friendly month labels
month_labels = {m: pd.to_datetime(m).strftime("%b %Y") for m in months}

# ── Disaggregated: month × location × product ────────────────────────────────
deagg = (
    all_data.groupby(["month_str","location","product"])
    .agg(net_sales=("net_sales","sum"), units=("net_sales","count"))
    .reset_index()
    .sort_values(["month_str","location","product"])
)

# ── Excel helpers ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")
ALT_FILL      = PatternFill("solid", fgColor="DCE6F1")
TOTAL_FILL    = PatternFill("solid", fgColor="BDD7EE")

H_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SH_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
B_FONT   = Font(bold=True, name="Calibri", size=10)
R_FONT   = Font(name="Calibri", size=10)

thin = Side(style="thin", color="B8B8B8")
BD   = Border(left=thin, right=thin, top=thin, bottom=thin)

NUM_FMT  = '#,##0.00'
INT_FMT  = '#,##0'

def style_header(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = HEADER_FILL
    c.font = H_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BD
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def style_subheader(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = SUBHDR_FILL
    c.font = SH_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BD
    return c

def style_cell(ws, row, col, value, fmt=None, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = B_FONT if bold else R_FONT
    c.border = BD
    c.alignment = Alignment(horizontal="right" if fmt else "left")
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: AGREGADO – Net Sales pivot
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Agregado_NetSales"
ws1.freeze_panes = "B3"

# Title
ws1.merge_cells("A1:S1")
title_cell = ws1.cell(row=1, column=1, value="FOOD ITEMS — NET SALES EVOLUTION (AGREGADO MENSUAL)")
title_cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
title_cell.fill = HEADER_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

# Headers
style_header(ws1, 2, 1, "Month", 12)
for ci, prod in enumerate(PRODUCTS, start=2):
    style_header(ws1, 2, ci, prod, 18)
style_header(ws1, 2, len(PRODUCTS)+2, "TOTAL", 14)

ws1.row_dimensions[2].height = 45

# Data rows
totals_row = [0.0] * len(PRODUCTS)
for ri, (month, row) in enumerate(pivot_sales.iterrows(), start=3):
    fill = ALT_FILL if ri % 2 == 0 else None
    style_cell(ws1, ri, 1, month_labels[month], bold=True, fill=fill)
    row_total = 0.0
    for ci, prod in enumerate(PRODUCTS, start=2):
        val = float(row[prod])
        style_cell(ws1, ri, ci, val if val != 0 else None, fmt=NUM_FMT, fill=fill)
        row_total += val
        totals_row[ci-2] += val
    style_cell(ws1, ri, len(PRODUCTS)+2, row_total, fmt=NUM_FMT, bold=True, fill=fill)

# Totals row
n_rows = len(months)
total_ri = n_rows + 3
style_cell(ws1, total_ri, 1, "TOTAL", bold=True, fill=TOTAL_FILL)
for ci, t in enumerate(totals_row, start=2):
    style_cell(ws1, total_ri, ci, t, fmt=NUM_FMT, bold=True, fill=TOTAL_FILL)
grand = sum(totals_row)
style_cell(ws1, total_ri, len(PRODUCTS)+2, grand, fmt=NUM_FMT, bold=True, fill=TOTAL_FILL)

# ── Line Chart: monthly net sales by product ──────────────────────────────────
chart_row_start = total_ri + 3
chart = LineChart()
chart.title = "Monthly Net Sales by Food Item"
chart.style = 10
chart.y_axis.title = "Net Sales (USD)"
chart.x_axis.title = "Month"
chart.height = 18
chart.width  = 38

# X-axis labels (months)
dates_ref = Reference(ws1, min_col=1, min_row=3, max_row=n_rows+2)
chart.set_categories(dates_ref)

# One series per product (skip products with all-zero sales)
for ci, prod in enumerate(PRODUCTS, start=2):
    col_vals = [pivot_sales[prod].iloc[ri] for ri in range(len(months))]
    if sum(abs(v) for v in col_vals) == 0:
        continue
    data_ref = Reference(ws1, min_col=ci, min_row=2, max_row=n_rows+2)
    chart.add_data(data_ref, titles_from_data=True)

ws1.add_chart(chart, f"A{chart_row_start}")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: AGREGADO – Units pivot
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Agregado_Unidades")
ws2.freeze_panes = "B3"

ws2.merge_cells("A1:S1")
tc2 = ws2.cell(row=1, column=1, value="FOOD ITEMS — UNITS SOLD EVOLUTION (AGREGADO MENSUAL)")
tc2.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
tc2.fill = HEADER_FILL
tc2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

style_header(ws2, 2, 1, "Month", 12)
for ci, prod in enumerate(PRODUCTS, start=2):
    style_header(ws2, 2, ci, prod, 18)
style_header(ws2, 2, len(PRODUCTS)+2, "TOTAL", 14)
ws2.row_dimensions[2].height = 45

u_totals = [0] * len(PRODUCTS)
for ri, (month, row) in enumerate(pivot_units.iterrows(), start=3):
    fill = ALT_FILL if ri % 2 == 0 else None
    style_cell(ws2, ri, 1, month_labels[month], bold=True, fill=fill)
    row_total = 0
    for ci, prod in enumerate(PRODUCTS, start=2):
        val = int(row[prod])
        style_cell(ws2, ri, ci, val if val != 0 else None, fmt=INT_FMT, fill=fill)
        row_total += val
        u_totals[ci-2] += val
    style_cell(ws2, ri, len(PRODUCTS)+2, row_total, fmt=INT_FMT, bold=True, fill=fill)

total_ri2 = n_rows + 3
style_cell(ws2, total_ri2, 1, "TOTAL", bold=True, fill=TOTAL_FILL)
for ci, t in enumerate(u_totals, start=2):
    style_cell(ws2, total_ri2, ci, t, fmt=INT_FMT, bold=True, fill=TOTAL_FILL)
style_cell(ws2, total_ri2, len(PRODUCTS)+2, sum(u_totals), fmt=INT_FMT, bold=True, fill=TOTAL_FILL)

# Bar chart: total units per product
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Total Units Sold per Food Item"
chart2.style = 10
chart2.y_axis.title = "Units"
chart2.height = 15
chart2.width  = 38

prod_labels = Reference(ws2, min_col=2, max_col=len(PRODUCTS)+1, min_row=2, max_row=2)
unit_totals_row = Reference(ws2, min_col=2, max_col=len(PRODUCTS)+1, min_row=total_ri2, max_row=total_ri2)
chart2.add_data(unit_totals_row)
chart2.set_categories(prod_labels)
chart2.series[0].title = SeriesLabel(v="Total Units")

chart2_start = total_ri2 + 3
ws2.add_chart(chart2, f"A{chart2_start}")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: DESAGREGADO – tabular
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Desagregado")
ws3.freeze_panes = "A3"

ws3.merge_cells("A1:F1")
tc3 = ws3.cell(row=1, column=1, value="FOOD ITEMS — NET SALES DESAGREGADO (MES × TIENDA × PRODUCTO)")
tc3.font = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
tc3.fill = HEADER_FILL
tc3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

headers3 = ["Month", "Location", "Product", "Net Sales", "Units", "Avg Ticket"]
col_widths3 = [12, 45, 35, 14, 10, 14]
for ci, (h, w) in enumerate(zip(headers3, col_widths3), start=1):
    style_header(ws3, 2, ci, h, w)
ws3.row_dimensions[2].height = 20

for ri, row in enumerate(deagg.itertuples(), start=3):
    fill = ALT_FILL if ri % 2 == 0 else None
    avg = row.net_sales / row.units if row.units > 0 else 0
    style_cell(ws3, ri, 1, month_labels.get(row.month_str, row.month_str), fill=fill)
    style_cell(ws3, ri, 2, row.location, fill=fill)
    style_cell(ws3, ri, 3, row.product, fill=fill)
    style_cell(ws3, ri, 4, row.net_sales, fmt=NUM_FMT, fill=fill)
    style_cell(ws3, ri, 5, row.units, fmt=INT_FMT, fill=fill)
    style_cell(ws3, ri, 6, avg, fmt=NUM_FMT, fill=fill)

# ── Bar chart: net sales by store (top products, all months combined) ──────────
store_prod = (
    all_data.groupby(["location","product"])["net_sales"]
    .sum()
    .reset_index()
    .sort_values("net_sales", ascending=False)
)

# Write store-product summary for chart to a helper area
chart_data_row = len(deagg) + 5
# Stores
stores = sorted(all_data["location"].unique())
ws3.cell(row=chart_data_row, column=1, value="Chart Data – Net Sales by Store")
ws3.cell(row=chart_data_row, column=1).font = B_FONT

style_cell(ws3, chart_data_row+1, 1, "Product", bold=True)
for si, store in enumerate(stores, start=2):
    short = store.replace("Qargo Coffee ", "").replace("**","").strip()
    ws3.cell(row=chart_data_row+1, column=si, value=short).font = B_FONT
    ws3.column_dimensions[get_column_letter(si)].width = 20

# products with any sales
active_prods = [p for p in PRODUCTS if all_data[all_data["product"]==p]["net_sales"].sum() > 0]
for pi, prod in enumerate(active_prods):
    ri = chart_data_row + 2 + pi
    ws3.cell(row=ri, column=1, value=prod).font = R_FONT
    for si, store in enumerate(stores, start=2):
        val = store_prod[(store_prod["location"]==store) & (store_prod["product"]==prod)]["net_sales"].sum()
        c = ws3.cell(row=ri, column=si, value=float(val) if val else None)
        c.number_format = NUM_FMT

chart3 = BarChart()
chart3.type = "bar"  # horizontal
chart3.title = "Net Sales by Store per Food Item"
chart3.style = 10
chart3.y_axis.title = "Food Item"
chart3.x_axis.title = "Net Sales (USD)"
chart3.height = 20
chart3.width  = 38

prod_labels3 = Reference(ws3, min_col=1, min_row=chart_data_row+2, max_row=chart_data_row+1+len(active_prods))
chart3.set_categories(prod_labels3)

for si, store in enumerate(stores, start=2):
    data_ref3 = Reference(ws3, min_col=si, min_row=chart_data_row+1, max_row=chart_data_row+1+len(active_prods))
    chart3.add_data(data_ref3, titles_from_data=True)

chart3_start = chart_data_row + len(active_prods) + 5
ws3.add_chart(chart3, f"A{chart3_start}")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: DATOS RAW
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Datos_Raw")
ws4.freeze_panes = "A2"

headers4 = ["Date", "Month", "Location", "Product", "Item Name (Raw)", "Net Sales", "Source"]
col_widths4 = [18, 12, 45, 35, 55, 14, 12]
for ci, (h, w) in enumerate(zip(headers4, col_widths4), start=1):
    style_header(ws4, 1, ci, h, w)

for ri, row in enumerate(all_data.sort_values(["month_str","location","product"]).itertuples(), start=2):
    fill = ALT_FILL if ri % 2 == 0 else None
    style_cell(ws4, ri, 1, row.date.strftime("%Y-%m-%d %H:%M") if pd.notna(row.date) else "", fill=fill)
    style_cell(ws4, ri, 2, month_labels.get(row.month_str,""), fill=fill)
    style_cell(ws4, ri, 3, row.location, fill=fill)
    style_cell(ws4, ri, 4, row.product, fill=fill)
    style_cell(ws4, ri, 5, row.item_name, fill=fill)
    style_cell(ws4, ri, 6, row.net_sales, fmt=NUM_FMT, fill=fill)
    style_cell(ws4, ri, 7, row.source, fill=fill)

# ── Save ───────────────────────────────────────────────────────────────────────
out = "food_sales_evolution.xlsx"
wb.save(out)
print(f"\nSaved → {out}")

# ── Summary ───────────────────────────────────────────────────────────────────
print("\n=== PRODUCT TOTALS (all time) ===")
summary = all_data.groupby("product").agg(
    net_sales=("net_sales","sum"),
    units=("net_sales","count")
).reindex(PRODUCTS).round(2)
summary["avg_ticket"] = (summary["net_sales"] / summary["units"]).round(2)
print(summary.to_string())
