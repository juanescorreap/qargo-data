"""
Tampa – "Open Item" sales | Oct 1 2025 – May 31 2026
Output: Excel with raw data, daily/weekly/monthly evolution, and charts
"""
import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = "/home/juan_esteban_correa/qargo-data/data"
OUTPUT   = "/home/juan_esteban_correa/qargo-data/tampa_open_item.xlsx"

FILES = [f"{DATA_DIR}/DDBB_{m}.csv"
         for m in ["Oct_25","Nov_25","Dec_25","Jan_26","Feb_26","Mar_26","Apr_26","May_26"]]

conn = duckdb.connect()

union_sql = " UNION ALL ".join([f"SELECT * FROM read_csv_auto('{f}')" for f in FILES])

MONEY = """CASE WHEN {col} LIKE '(%)'
    THEN -CAST(REPLACE(REPLACE(REPLACE(REPLACE({col},'$',''),',',''),'(',''),')','') AS DOUBLE)
    ELSE  CAST(REPLACE(REPLACE(REPLACE(REPLACE({col},'$',''),',',''),'(',''),')','') AS DOUBLE)
END"""

raw_df = conn.execute(f"""
    SELECT
        Location,
        \"Closed Date/Time\"                                                AS closed_datetime,
        strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p')                AS ts,
        CAST(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p') AS DATE)  AS sale_date,
        DATE_TRUNC('week', strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p'))::DATE AS week_start,
        MONTHNAME(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p'))     AS month_name,
        MONTH(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p'))          AS month_num,
        YEAR(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p'))           AS year_num,
        \"Employee Name\"   AS employee,
        \"Item Name\"       AS item_name,
        \"Revenue Center\"  AS revenue_center,
        \"Destination\"     AS destination,
        \"Order ID\"        AS order_id,
        {MONEY.format(col='"Net Sales"')}       AS net_sales,
        {MONEY.format(col='"Discount Total"')}  AS discount_total,
        {MONEY.format(col='"Gross Sales"')}     AS gross_sales,
        \"Voided\"          AS voided
    FROM ({union_sql})
    WHERE Location LIKE '%Tampa%'
    AND   lower(\"Item Name\") = 'open item'
    AND   strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p') >= '2025-10-01'
    AND   strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p') <  '2026-06-01'
    ORDER BY ts
""").df()

print(f"Rows: {len(raw_df)}")
print(f"Net sales total: ${raw_df['net_sales'].sum():,.2f}")
print(raw_df.groupby('month_name')['net_sales'].sum())

# Daily aggregation (exclude voided)
daily_df = conn.execute("""
    SELECT
        sale_date,
        COUNT(*)                      AS transactions,
        COUNT(DISTINCT order_id)      AS orders,
        ROUND(SUM(net_sales), 2)      AS net_sales,
        ROUND(AVG(net_sales), 2)      AS avg_net_sales,
        ROUND(SUM(gross_sales), 2)    AS gross_sales,
        ROUND(SUM(discount_total), 2) AS total_discounts
    FROM raw_df
    WHERE voided = 'False'
    GROUP BY sale_date
    ORDER BY sale_date
""").df()

# Weekly aggregation (exclude voided)
weekly_df = conn.execute("""
    SELECT
        week_start,
        COUNT(*)                      AS transactions,
        COUNT(DISTINCT order_id)      AS orders,
        ROUND(SUM(net_sales), 2)      AS net_sales,
        ROUND(SUM(gross_sales), 2)    AS gross_sales,
        ROUND(SUM(discount_total), 2) AS total_discounts
    FROM raw_df
    WHERE voided = 'False'
    GROUP BY week_start
    ORDER BY week_start
""").df()

# Monthly aggregation (exclude voided)
monthly_df = conn.execute("""
    SELECT
        year_num || '-' || LPAD(CAST(month_num AS VARCHAR),2,'0') AS year_month,
        month_name,
        COUNT(*)                      AS transactions,
        COUNT(DISTINCT order_id)      AS orders,
        ROUND(SUM(net_sales), 2)      AS net_sales,
        ROUND(SUM(gross_sales), 2)    AS gross_sales,
        ROUND(SUM(discount_total), 2) AS total_discounts
    FROM raw_df
    WHERE voided = 'False'
    GROUP BY year_num, month_num, month_name
    ORDER BY year_num, month_num
""").df()

# By employee (exclude voided)
employee_df = conn.execute("""
    SELECT
        employee,
        COUNT(*)               AS transactions,
        ROUND(SUM(net_sales), 2) AS net_sales,
        ROUND(AVG(net_sales), 2) AS avg_net_sales
    FROM raw_df
    WHERE voided = 'False'
    GROUP BY employee
    ORDER BY net_sales DESC
""").df()

# By destination (exclude voided)
dest_df = conn.execute("""
    SELECT
        destination,
        COUNT(*)               AS transactions,
        ROUND(SUM(net_sales), 2) AS net_sales
    FROM raw_df
    WHERE voided = 'False'
    GROUP BY destination
    ORDER BY net_sales DESC
""").df()

# ── Excel ──────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="922B21")   # dark red (Tampa brand)
SUB_FILL  = PatternFill("solid", fgColor="C0392B")
ALT_FILL  = PatternFill("solid", fgColor="FDEDEC")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(bold=True, name="Calibri", size=10)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")
thin = Border(
    left=Side(style="thin", color="BDC3C7"), right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),  bottom=Side(style="thin", color="BDC3C7"),
)

def title_row(ws, text, merge_to, color="922B21"):
    ws["A1"].value = text
    ws["A1"].font  = Font(bold=True, size=13, name="Calibri", color=color)
    ws.merge_cells(f"A1:{merge_to}1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CENTER; cell.border = thin

def write_df(ws, df, start_row=2):
    for ci, col in enumerate(df.columns, 1):
        ws.cell(start_row, ci, str(col).replace("_"," ").title())
    style_header(ws, start_row, len(df.columns))
    for ri, (_, row) in enumerate(df.iterrows()):
        fill = ALT_FILL if ri % 2 else PatternFill()
        for ci, val in enumerate(row, 1):
            c = ws.cell(start_row+1+ri, ci, val)
            c.font = BODY_FONT; c.border = thin; c.fill = fill
            c.alignment = CENTER if isinstance(val, (int,float)) else LEFT
    return start_row + 1 + len(df)

wb = Workbook()

# ── Sheet 1: Raw Data ──────────────────────────────────────────────────────────
ws_raw = wb.active
ws_raw.title = "Raw Data"
title_row(ws_raw, "Qargo Coffee Tampa, FL – Open Item Transactions | Oct 2025 – May 2026", "P")
export_cols = ["closed_datetime","sale_date","month_name","employee","item_name",
               "revenue_center","destination","order_id","net_sales","gross_sales","discount_total","voided"]
write_df(ws_raw, raw_df[export_cols])
widths = [22,12,10,24,12,14,14,18,12,13,16]
for i,w in enumerate(widths,1):
    ws_raw.column_dimensions[get_column_letter(i)].width = w
ws_raw.freeze_panes = "A3"

# ── Sheet 2: Monthly ──────────────────────────────────────────────────────────
ws_m = wb.create_sheet("Monthly")
title_row(ws_m, "Qargo Coffee Tampa, FL – Open Item | Monthly Summary", "G")
end = write_df(ws_m, monthly_df)
# totals
totals_row = end
for ci,val in enumerate([
    "TOTAL", "", int(monthly_df.transactions.sum()), int(monthly_df.orders.sum()),
    round(float(monthly_df.net_sales.sum()),2),
    round(float(monthly_df.gross_sales.sum()),2),
    round(float(monthly_df.total_discounts.sum()),2)
], 1):
    c = ws_m.cell(totals_row, ci, val)
    c.fill = SUB_FILL; c.font = Font(bold=True,color="FFFFFF",name="Calibri")
    c.border = thin; c.alignment = CENTER
for i,w in enumerate([14,12,13,10,13,14,16],1):
    ws_m.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 3: Daily ─────────────────────────────────────────────────────────────
ws_d = wb.create_sheet("Daily")
title_row(ws_d, "Qargo Coffee Tampa, FL – Open Item | Daily Summary", "G")
write_df(ws_d, daily_df)
for i,w in enumerate([14,13,10,13,15,14,16],1):
    ws_d.column_dimensions[get_column_letter(i)].width = w
ws_d.freeze_panes = "A3"

# ── Sheet 4: Weekly ────────────────────────────────────────────────────────────
ws_w = wb.create_sheet("Weekly")
title_row(ws_w, "Qargo Coffee Tampa, FL – Open Item | Weekly Summary", "F")
write_df(ws_w, weekly_df)
for i,w in enumerate([14,13,10,13,14,16],1):
    ws_w.column_dimensions[get_column_letter(i)].width = w
ws_w.freeze_panes = "A3"

# ── Sheet 5: By Employee ───────────────────────────────────────────────────────
ws_e = wb.create_sheet("By Employee")
title_row(ws_e, "Qargo Coffee Tampa, FL – Open Item | By Employee", "D")
write_df(ws_e, employee_df)
for i,w in enumerate([28,14,13,15],1):
    ws_e.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 6: Charts ────────────────────────────────────────────────────────────
ws_c = wb.create_sheet("Charts")
title_row(ws_c, "Qargo Coffee Tampa, FL – Open Item | Sales Evolution Oct 2025 – May 2026", "Z")

# Write source data for charts
# Monthly data cols: A=year_month, B=transactions, C=orders, D=net_sales
ws_c["A3"].value = "Month"
ws_c["B3"].value = "Net Sales ($)"
ws_c["C3"].value = "Transactions"
ws_c["D3"].value = "Orders"
for cell in [ws_c["A3"], ws_c["B3"], ws_c["C3"], ws_c["D3"]]:
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER

for i, row in monthly_df.iterrows():
    r = i + 4
    ws_c.cell(r, 1, row["year_month"])
    ws_c.cell(r, 2, float(row["net_sales"]))
    ws_c.cell(r, 3, int(row["transactions"]))
    ws_c.cell(r, 4, int(row["orders"]))

n_m = len(monthly_df)

# Write weekly data starting at col F
ws_c["F3"].value = "Week Start"
ws_c["G3"].value = "Net Sales ($)"
ws_c["H3"].value = "Orders"
for cell in [ws_c["F3"], ws_c["G3"], ws_c["H3"]]:
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER

for i, row in weekly_df.iterrows():
    r = i + 4
    ws_c.cell(r, 6, str(row["week_start"]))
    ws_c.cell(r, 7, float(row["net_sales"]))
    ws_c.cell(r, 8, int(row["orders"]))

n_w = len(weekly_df)

# Chart 1 — Monthly Net Sales (Line)
chart1 = LineChart()
chart1.title = "Monthly Net Sales – Open Item (Tampa)"
chart1.y_axis.title = "Net Sales (USD)"
chart1.x_axis.title = "Month"
chart1.style = 10; chart1.width = 26; chart1.height = 14

d1 = Reference(ws_c, min_col=2, max_col=2, min_row=3, max_row=3+n_m)
cats1 = Reference(ws_c, min_col=1, min_row=4, max_row=3+n_m)
chart1.add_data(d1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.line.solidFill = "922B21"
chart1.series[0].graphicalProperties.line.width = 25000
chart1.series[0].marker.symbol = "circle"
chart1.series[0].marker.size = 7
ws_c.add_chart(chart1, "A5")

# Chart 2 — Monthly Transactions Bar
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Monthly Transactions – Open Item (Tampa)"
chart2.y_axis.title = "Transactions"
chart2.x_axis.title = "Month"
chart2.style = 10; chart2.width = 26; chart2.height = 14

d2 = Reference(ws_c, min_col=3, max_col=3, min_row=3, max_row=3+n_m)
chart2.add_data(d2, titles_from_data=True)
chart2.set_categories(cats1)
chart2.series[0].graphicalProperties.solidFill = "C0392B"
chart2.series[0].graphicalProperties.line.solidFill = "922B21"
ws_c.add_chart(chart2, "A28")

# Chart 3 — Weekly Net Sales Line
chart3 = LineChart()
chart3.title = "Weekly Net Sales – Open Item (Tampa)"
chart3.y_axis.title = "Net Sales (USD)"
chart3.x_axis.title = "Week"
chart3.style = 10; chart3.width = 36; chart3.height = 14

d3 = Reference(ws_c, min_col=7, max_col=7, min_row=3, max_row=3+n_w)
cats3 = Reference(ws_c, min_col=6, min_row=4, max_row=3+n_w)
chart3.add_data(d3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.series[0].graphicalProperties.line.solidFill = "5B2C6F"
chart3.series[0].graphicalProperties.line.width = 20000
chart3.series[0].marker.symbol = "diamond"
chart3.series[0].marker.size = 5
ws_c.add_chart(chart3, "A51")

for i,w in enumerate([14,14,14,12,2,14,14,12],1):
    ws_c.column_dimensions[get_column_letter(i)].width = w

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Sheets: Raw Data ({len(raw_df)} rows) | Monthly | Daily | Weekly | By Employee | Charts")
