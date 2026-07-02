"""
Westerville hourly sales - last 3 months (Mar-May 2026)
Output: Excel with raw data, summary table, and chart
"""
import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

DATA_DIR = "/home/juan_esteban_correa/qargo-data/data"
OUTPUT = "/home/juan_esteban_correa/qargo-data/westerville_hourly_sales.xlsx"
STORE = "%Westerville%"
FILES = [
    f"{DATA_DIR}/DDBB_Mar_26.csv",
    f"{DATA_DIR}/DDBB_Apr_26.csv",
    f"{DATA_DIR}/DDBB_May_26.csv",
]

conn = duckdb.connect()

union_sql = " UNION ALL ".join([
    f"SELECT * FROM read_csv_auto('{f}')" for f in FILES
])

raw_df = conn.execute(f"""
    SELECT
        Location,
        \"Closed Date/Time\" AS closed_datetime,
        \"Employee Name\" AS employee,
        \"Item Name\" AS item_name,
        \"Revenue Center\" AS revenue_center,
        \"Destination\" AS destination,
        \"Order ID\" AS order_id,
        CASE
            WHEN \"Net Sales\" LIKE '(%)'
            THEN -CAST(REPLACE(REPLACE(REPLACE(REPLACE(\"Net Sales\", '$', ''), ',', ''), '(', ''), ')', '') AS DOUBLE)
            ELSE CAST(REPLACE(REPLACE(REPLACE(REPLACE(\"Net Sales\", '$', ''), ',', ''), '(', ''), ')', '') AS DOUBLE)
        END AS net_sales,
        CASE
            WHEN \"Discount Total\" LIKE '(%)'
            THEN -CAST(REPLACE(REPLACE(REPLACE(REPLACE(\"Discount Total\", '$', ''), ',', ''), '(', ''), ')', '') AS DOUBLE)
            ELSE CAST(REPLACE(REPLACE(REPLACE(REPLACE(\"Discount Total\", '$', ''), ',', ''), '(', ''), ')', '') AS DOUBLE)
        END AS discount_total,
        \"Voided\" AS voided,
        \"Is Modifier\" AS is_modifier,
        strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p') AS ts,
        HOUR(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p')) AS hour_of_day,
        DAYNAME(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p')) AS day_of_week,
        MONTHNAME(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p')) AS month_name,
        CAST(strptime(\"Closed Date/Time\", '%m/%d/%Y %I:%M %p') AS DATE) AS sale_date
    FROM ({union_sql})
    WHERE Location LIKE '{STORE}'
    AND Voided = 'False'
    ORDER BY ts
""").df()

print(f"Raw rows (non-voided): {len(raw_df)}")

# Hourly summary - aggregate at order+item level (net_sales already per line item)
hourly_df = conn.execute("""
    SELECT
        hour_of_day                                     AS hour,
        LPAD(CAST(hour_of_day AS VARCHAR), 2, '0') || ':00'  AS hour_label,
        COUNT(DISTINCT order_id)                        AS orders,
        ROUND(SUM(net_sales), 2)                        AS net_sales,
        ROUND(SUM(net_sales) / COUNT(DISTINCT order_id), 2) AS avg_ticket,
        ROUND(100.0 * SUM(net_sales) / SUM(SUM(net_sales)) OVER (), 2) AS pct_of_sales
    FROM raw_df
    WHERE is_modifier = 'False'
    GROUP BY hour_of_day
    ORDER BY hour_of_day
""").df()

# Monthly hourly breakdown
monthly_hourly_df = conn.execute("""
    SELECT
        month_name                                      AS month,
        hour_of_day                                     AS hour,
        LPAD(CAST(hour_of_day AS VARCHAR), 2, '0') || ':00'  AS hour_label,
        COUNT(DISTINCT order_id)                        AS orders,
        ROUND(SUM(net_sales), 2)                        AS net_sales
    FROM raw_df
    WHERE is_modifier = 'False'
    GROUP BY month_name, hour_of_day
    ORDER BY hour_of_day,
        CASE month_name WHEN 'March' THEN 1 WHEN 'April' THEN 2 WHEN 'May' THEN 3 END
""").df()

# Daily summary
daily_df = conn.execute("""
    SELECT
        sale_date,
        day_of_week,
        COUNT(DISTINCT order_id) AS orders,
        ROUND(SUM(net_sales), 2) AS net_sales
    FROM raw_df
    WHERE is_modifier = 'False'
    GROUP BY sale_date, day_of_week
    ORDER BY sale_date
""").df()

print("Hourly summary:")
print(hourly_df.to_string())

# ── Excel ──────────────────────────────────────────────────────────────────────
wb = Workbook()

# Colors
HDR_FILL  = PatternFill("solid", fgColor="1B3A5C")   # dark blue
SUB_FILL  = PatternFill("solid", fgColor="2E86C1")   # mid blue
ALT_FILL  = PatternFill("solid", fgColor="EBF5FB")   # light blue row
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(bold=True, name="Calibri", size=10)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")

thin_border = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = thin_border

def write_df_to_sheet(ws, df, start_row=1, header=True):
    if header:
        for ci, col in enumerate(df.columns, 1):
            ws.cell(row=start_row, column=ci, value=str(col).replace("_", " ").title())
        style_header_row(ws, start_row, len(df.columns))
        start_row += 1
    for ri, (_, row) in enumerate(df.iterrows()):
        fill = ALT_FILL if ri % 2 == 1 else PatternFill()
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=start_row + ri, column=ci, value=val)
            c.font = BODY_FONT
            c.border = thin_border
            c.fill = fill
            if isinstance(val, float):
                c.alignment = CENTER
            else:
                c.alignment = LEFT
    return start_row + len(df)

# ── Sheet 1: Raw Data ──────────────────────────────────────────────────────────
ws_raw = wb.active
ws_raw.title = "Raw Data"

title_cell = ws_raw["A1"]
title_cell.value = "Qargo Coffee – Westerville, OH | Raw Transactions | Mar–May 2026"
title_cell.font = Font(bold=True, size=13, name="Calibri", color="1B3A5C")
ws_raw.merge_cells("A1:N1")
ws_raw["A1"].alignment = CENTER
ws_raw.row_dimensions[1].height = 28

export_cols = ["closed_datetime","hour_of_day","day_of_week","month_name",
               "sale_date","employee","item_name","revenue_center",
               "destination","order_id","net_sales","discount_total","is_modifier"]
raw_export = raw_df[export_cols].copy()
write_df_to_sheet(ws_raw, raw_export, start_row=2)

col_widths = [22,12,14,10,12,24,32,14,14,18,12,14,12]
for i, w in enumerate(col_widths, 1):
    ws_raw.column_dimensions[get_column_letter(i)].width = w

ws_raw.freeze_panes = "A3"

# ── Sheet 2: Hourly Summary ────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Hourly Summary")

ws_sum["A1"].value = "Qargo Coffee – Westerville, OH | Hourly Sales Summary | Mar–May 2026"
ws_sum["A1"].font = Font(bold=True, size=13, name="Calibri", color="1B3A5C")
ws_sum.merge_cells("A1:F1")
ws_sum["A1"].alignment = CENTER
ws_sum.row_dimensions[1].height = 28

end_row = write_df_to_sheet(ws_sum, hourly_df, start_row=2)

# Totals row
totals_row = end_row
ws_sum.cell(totals_row, 1, "TOTAL").font = BOLD_FONT
ws_sum.cell(totals_row, 1).fill = SUB_FILL
ws_sum.cell(totals_row, 1).font = Font(bold=True, color="FFFFFF", name="Calibri")
ws_sum.cell(totals_row, 3, int(hourly_df["orders"].sum()))
ws_sum.cell(totals_row, 4, round(float(hourly_df["net_sales"].sum()), 2))
ws_sum.cell(totals_row, 5, round(float(hourly_df["net_sales"].sum()) / int(hourly_df["orders"].sum()), 2))
ws_sum.cell(totals_row, 6, 100.0)
for col in range(1, 7):
    c = ws_sum.cell(totals_row, col)
    c.fill = SUB_FILL
    c.font = Font(bold=True, color="FFFFFF", name="Calibri")
    c.border = thin_border
    c.alignment = CENTER

for i, w in enumerate([10, 12, 10, 14, 13, 14], 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = w

ws_sum.freeze_panes = "A3"

# ── Sheet 3: Monthly Breakdown ─────────────────────────────────────────────────
ws_monthly = wb.create_sheet("Monthly Breakdown")
ws_monthly["A1"].value = "Qargo Coffee – Westerville, OH | Hourly Sales by Month | Mar–May 2026"
ws_monthly["A1"].font = Font(bold=True, size=13, name="Calibri", color="1B3A5C")
ws_monthly.merge_cells("A1:E1")
ws_monthly["A1"].alignment = CENTER
ws_monthly.row_dimensions[1].height = 28

write_df_to_sheet(ws_monthly, monthly_hourly_df, start_row=2)
for i, w in enumerate([12, 8, 12, 10, 13], 1):
    ws_monthly.column_dimensions[get_column_letter(i)].width = w
ws_monthly.freeze_panes = "A3"

# ── Sheet 4: Daily Summary ─────────────────────────────────────────────────────
ws_daily = wb.create_sheet("Daily Summary")
ws_daily["A1"].value = "Qargo Coffee – Westerville, OH | Daily Sales | Mar–May 2026"
ws_daily["A1"].font = Font(bold=True, size=13, name="Calibri", color="1B3A5C")
ws_daily.merge_cells("A1:D1")
ws_daily["A1"].alignment = CENTER
ws_daily.row_dimensions[1].height = 28
write_df_to_sheet(ws_daily, daily_df, start_row=2)
for i, w in enumerate([14, 14, 10, 13], 1):
    ws_daily.column_dimensions[get_column_letter(i)].width = w
ws_daily.freeze_panes = "A3"

# ── Sheet 5: Chart ─────────────────────────────────────────────────────────────
ws_chart = wb.create_sheet("Chart")
ws_chart["A1"].value = "Qargo Coffee – Westerville, OH | Net Sales by Hour of Day | Mar–May 2026"
ws_chart["A1"].font = Font(bold=True, size=13, name="Calibri", color="1B3A5C")
ws_chart.merge_cells("A1:P1")
ws_chart["A1"].alignment = CENTER
ws_chart.row_dimensions[1].height = 28

# Write mini data table for chart source
ws_chart["A3"].value = "Hour"
ws_chart["B3"].value = "Net Sales ($)"
ws_chart["C3"].value = "Orders"
ws_chart["A3"].font = HDR_FONT; ws_chart["A3"].fill = HDR_FILL; ws_chart["A3"].alignment = CENTER
ws_chart["B3"].font = HDR_FONT; ws_chart["B3"].fill = HDR_FILL; ws_chart["B3"].alignment = CENTER
ws_chart["C3"].font = HDR_FONT; ws_chart["C3"].fill = HDR_FILL; ws_chart["C3"].alignment = CENTER

for i, row in hourly_df.iterrows():
    r = i + 4
    ws_chart.cell(r, 1, row["hour_label"])
    ws_chart.cell(r, 2, float(row["net_sales"]))
    ws_chart.cell(r, 3, int(row["orders"]))

n_rows = len(hourly_df)

# Bar chart - Net Sales
chart = BarChart()
chart.type = "col"
chart.title = "Net Sales by Hour of Day – Westerville (Mar–May 2026)"
chart.y_axis.title = "Net Sales (USD)"
chart.x_axis.title = "Hour of Day"
chart.style = 10
chart.width = 28
chart.height = 16

data = Reference(ws_chart, min_col=2, max_col=2, min_row=3, max_row=3 + n_rows)
cats = Reference(ws_chart, min_col=1, min_row=4, max_row=3 + n_rows)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = "2E86C1"
chart.series[0].graphicalProperties.line.solidFill = "1B3A5C"

ws_chart.add_chart(chart, "A5")

# Orders bar chart below
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Order Count by Hour of Day – Westerville (Mar–May 2026)"
chart2.y_axis.title = "Orders"
chart2.x_axis.title = "Hour of Day"
chart2.style = 10
chart2.width = 28
chart2.height = 14

data2 = Reference(ws_chart, min_col=3, max_col=3, min_row=3, max_row=3 + n_rows)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats)
chart2.series[0].graphicalProperties.solidFill = "1ABC9C"
chart2.series[0].graphicalProperties.line.solidFill = "17A589"

ws_chart.add_chart(chart2, "A36")

ws_chart.column_dimensions["A"].width = 12
ws_chart.column_dimensions["B"].width = 16
ws_chart.column_dimensions["C"].width = 12

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Sheets: Raw Data ({len(raw_export)} rows) | Hourly Summary | Monthly Breakdown | Daily Summary | Chart")
